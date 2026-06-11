#!/usr/bin/env python3
"""
Pre-aggregate official DOL OFLC LCA Disclosure Data into the compact, nested
JSON consumed by the interactive /tools/h1b-salaries explorer.

Raw LCA rows NEVER ship to the browser — only per-(role, metro, level) and
coarser percentile aggregates, histograms, and top-employer medians.

Streaming: the source .xlsx is ~138MB / ~1M rows. We read it row-by-row with
openpyxl read_only mode (constant memory for the parse) and accumulate only
the numbers we need.

Usage:
    pip3 install openpyxl
    python3 scripts/build-h1b-data.py "/path/to/LCA_Disclosure_Data_FY2026_Q2.xlsx"
    # CSV also accepted (same columns):
    python3 scripts/build-h1b-data.py data.csv

Output: public/data/h1b/explorer.json  (+ prints sanity checks)
Refresh cadence + rules: docs/DATA-UPDATE-PLAYBOOK.md
"""
import sys, os, json, csv, math
from collections import defaultdict

PERIOD = "FY2026 Q2"
SOURCE = "DOL OFLC LCA Disclosure Data"
SOURCE_URL = "https://www.dol.gov/agencies/eta/foreign-labor/performance"

NEEDED = [
    "CASE_STATUS", "VISA_CLASS", "JOB_TITLE", "SOC_CODE", "SOC_TITLE",
    "FULL_TIME_POSITION", "EMPLOYER_NAME", "WORKSITE_CITY", "WORKSITE_STATE",
    "WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY_TO", "WAGE_UNIT_OF_PAY",
    "PREVAILING_WAGE", "PW_UNIT_OF_PAY", "PW_WAGE_LEVEL",
]

WAGE_MULT = {"YEAR": 1, "MONTH": 12, "BI-WEEKLY": 26, "WEEK": 52, "HOUR": 2080}
LEVELS = {"I": 1, "II": 2, "III": 3, "IV": 4}

# Output thresholds — tuned to keep the JSON to a few MB while covering the
# selections users actually make.
MIN_ROLE = 40          # drop ultra-rare occupations entirely
MIN_METRO = 10         # (role, metro) cell must have this many to publish
MIN_METRO_LEVEL = 5    # (role, metro, level) cell
MIN_STATE = 10         # (role, state)
MIN_LEVEL = 10         # (role, level)
MIN_EMP = 5            # (role, employer)
MAX_METROS_PER_ROLE = 80
MAX_EMP_PER_ROLE = 15
HIST_BUCKET = 20000    # $20k histogram buckets
HIST_MAX = 400000      # everything >= this lands in the top bucket

LEGAL_SUFFIX = {
    "INC", "LLC", "LTD", "CORP", "CORPORATION", "CO", "LP", "LLP", "PLLC",
    "PC", "LIMITED", "INCORPORATED", "L.L.C", "L.P", "COMPANY",
}


def to_annual(val, unit):
    try:
        v = float(val)
    except (TypeError, ValueError):
        return None
    if v <= 0:
        return None
    m = WAGE_MULT.get(str(unit or "").strip().upper())
    if not m:
        return None
    a = v * m
    if a < 15000 or a > 5_000_000:
        return None
    return a


def norm_employer(raw):
    if not raw:
        return None, None
    disp = " ".join(str(raw).split()).strip().strip(".,")
    if not disp:
        return None, None
    key = disp.upper().replace(".", "")
    toks = key.split()
    while len(toks) > 1 and toks[-1] in LEGAL_SUFFIX:
        toks.pop()
    key = " ".join(toks) if toks else key
    return key, disp


def title_city(c):
    return " ".join(w.capitalize() for w in str(c).split())


def rows_from_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    idx = {h: i for i, h in enumerate(header)}
    for n in NEEDED:
        if n not in idx:
            raise SystemExit(f"Column {n} not found in {path}")
    cols = {n: idx[n] for n in NEEDED}
    for r in it:
        yield {n: r[cols[n]] for n in NEEDED}


def rows_from_csv(path):
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield row


def pctile(sorted_vals, p):
    if not sorted_vals:
        return 0
    if len(sorted_vals) == 1:
        return sorted_vals[0]
    idx = (len(sorted_vals) - 1) * p
    lo, hi = math.floor(idx), math.ceil(idx)
    v = sorted_vals[lo] + (sorted_vals[hi] - sorted_vals[lo]) * (idx - lo)
    return v


def r1000(x):
    return int(round(x / 1000.0) * 1000)


def stats(vals, pw_vals=None):
    s = sorted(vals)
    out = {
        "n": len(s),
        "min": r1000(s[0]),
        "p10": r1000(pctile(s, 0.10)),
        "p25": r1000(pctile(s, 0.25)),
        "p50": r1000(pctile(s, 0.50)),
        "p75": r1000(pctile(s, 0.75)),
        "p90": r1000(pctile(s, 0.90)),
        "max": r1000(s[-1]),
    }
    if pw_vals:
        out["medPW"] = r1000(pctile(sorted(pw_vals), 0.50))
    return out


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python3 scripts/build-h1b-data.py <file.xlsx|file.csv>")
    path = sys.argv[1]
    here = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(here, "..", "public", "data", "h1b")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "explorer.json")

    reader = rows_from_csv(path) if path.lower().endswith(".csv") else rows_from_xlsx(path)

    # Accumulators
    sal_rml = defaultdict(list)  # (role, metro, level) -> [salary]
    pw_rml = defaultdict(list)
    sal_rm = defaultdict(list)   # (role, metro) -> [salary]
    pw_rm = defaultdict(list)
    sal_rs = defaultdict(list)   # (role, state) -> [salary]
    sal_rl = defaultdict(list)   # (role, level) -> [salary]
    sal_emp = defaultdict(lambda: defaultdict(list))  # role -> empkey -> [salary]
    emp_disp = {}
    role_count = defaultdict(int)
    metro_count = defaultdict(int)
    state_count = defaultdict(int)
    level_count = defaultdict(int)

    processed = kept = withdrawn = 0
    for row in reader:
        processed += 1
        if processed % 200000 == 0:
            print(f"  ...{processed:,} rows scanned, {kept:,} kept", file=sys.stderr)

        if str(row.get("VISA_CLASS") or "").strip().upper() != "H-1B":
            continue
        status = str(row.get("CASE_STATUS") or "").strip()
        if not status.lower().startswith("certified"):
            continue
        is_withdrawn = "withdrawn" in status.lower()

        role = str(row.get("SOC_TITLE") or "").strip()
        if not role:
            continue
        salary = to_annual(row.get("WAGE_RATE_OF_PAY_FROM"), row.get("WAGE_UNIT_OF_PAY"))
        if salary is None:
            continue
        pw = to_annual(row.get("PREVAILING_WAGE"), row.get("PW_UNIT_OF_PAY"))

        city = str(row.get("WORKSITE_CITY") or "").strip()
        state = str(row.get("WORKSITE_STATE") or "").strip().upper()
        metro = f"{title_city(city)}, {state}" if city and state else None
        level = LEVELS.get(str(row.get("PW_WAGE_LEVEL") or "").strip().upper())

        kept += 1
        if is_withdrawn:
            withdrawn += 1
        role_count[role] += 1

        if metro:
            sal_rm[(role, metro)].append(salary)
            if pw:
                pw_rm[(role, metro)].append(pw)
            metro_count[metro] += 1
            if level:
                sal_rml[(role, metro, level)].append(salary)
                if pw:
                    pw_rml[(role, metro, level)].append(pw)
        if state:
            sal_rs[(role, state)].append(salary)
            state_count[state] += 1
        if level:
            sal_rl[(role, level)].append(salary)
            level_count[level] += 1

        empkey, disp = norm_employer(row.get("EMPLOYER_NAME"))
        if empkey:
            sal_emp[role][empkey].append(salary)
            if empkey not in emp_disp:
                emp_disp[empkey] = disp

    print(f"Scanned {processed:,} rows; kept {kept:,} certified H-1B "
          f"({withdrawn:,} certified-withdrawn).", file=sys.stderr)

    # Build per-role nested output
    roles_out = {}
    kept_roles = [r for r, c in role_count.items() if c >= MIN_ROLE]
    for role in kept_roles:
        # metros for this role
        metros = [(m, sal_rm[(role, m)]) for (rr, m) in sal_rm if rr == role]
        metros = [(m, v) for m, v in metros if len(v) >= MIN_METRO]
        metros.sort(key=lambda mv: len(mv[1]), reverse=True)
        metros = metros[:MAX_METROS_PER_ROLE]

        by_metro = {}
        by_metro_level = {}
        for m, v in metros:
            st = stats(v, pw_rm.get((role, m)))
            # histogram
            hist = defaultdict(int)
            for s in v:
                b = min(int(s // HIST_BUCKET) * HIST_BUCKET, HIST_MAX)
                hist[b] += 1
            st["hist"] = [{"b": b, "c": hist[b]} for b in sorted(hist)]
            by_metro[m] = st
            for lvl in (1, 2, 3, 4):
                vv = sal_rml.get((role, m, lvl))
                if vv and len(vv) >= MIN_METRO_LEVEL:
                    by_metro_level[f"{m}|{lvl}"] = stats(vv, pw_rml.get((role, m, lvl)))

        by_state = {}
        for (rr, stt) in sal_rs:
            if rr != role:
                continue
            v = sal_rs[(rr, stt)]
            if len(v) >= MIN_STATE:
                by_state[stt] = stats(v)

        by_level = {}
        for lvl in (1, 2, 3, 4):
            v = sal_rl.get((role, lvl))
            if v and len(v) >= MIN_LEVEL:
                by_level[str(lvl)] = stats(v)

        emps = []
        for ek, v in sal_emp.get(role, {}).items():
            if len(v) >= MIN_EMP:
                emps.append((emp_disp[ek], len(v), r1000(pctile(sorted(v), 0.5))))
        emps.sort(key=lambda e: e[1], reverse=True)
        emps = [{"name": n, "n": c, "p50": med} for n, c, med in emps[:MAX_EMP_PER_ROLE]]

        all_role_vals = sorted(
            s for (rr, m) in sal_rm if rr == role for s in sal_rm[(rr, m)]
        )
        roles_out[role] = {
            "count": role_count[role],
            "medAll": r1000(pctile(all_role_vals, 0.5)) if all_role_vals else 0,
            "byMetro": by_metro,
            "byMetroLevel": by_metro_level,
            "byState": by_state,
            "byLevel": by_level,
            "employers": emps,
        }

    # Filters index
    roles_idx = sorted(
        ({"key": r, "count": role_count[r]} for r in roles_out),
        key=lambda x: x["count"], reverse=True,
    )
    metros_idx = sorted(
        ({"key": m, "count": c} for m, c in metro_count.items() if c >= MIN_METRO * 3),
        key=lambda x: x["count"], reverse=True,
    )
    states_idx = sorted(
        ({"key": s, "count": c} for s, c in state_count.items()),
        key=lambda x: x["count"], reverse=True,
    )
    levels_idx = [{"level": l, "count": level_count.get(l, 0)} for l in (1, 2, 3, 4)]

    from datetime import date
    out = {
        "lastUpdated": date.today().isoformat(),
        "source": SOURCE,
        "sourceUrl": SOURCE_URL,
        "period": PERIOD,
        "visaClass": "H-1B",
        "recordsProcessed": processed,
        "recordsKept": kept,
        "withdrawnCount": withdrawn,
        "filters": {
            "roles": roles_idx,
            "metros": metros_idx,
            "states": states_idx,
            "levels": levels_idx,
        },
        "roles": roles_out,
    }

    with open(out_path, "w") as f:
        json.dump(out, f, separators=(",", ":"))
    size_mb = os.path.getsize(out_path) / 1e6
    print(f"\nWrote {out_path}  ({size_mb:.2f} MB, {len(roles_out)} roles)")

    # ---- Sanity checks ----
    print("\nSANITY — Software Developers median by metro+level:")
    sd = roles_out.get("Software Developers", {})
    samples = []
    for m in ["San Jose, CA", "Seattle, WA", "New York, NY", "Austin, TX", "San Francisco, CA"]:
        st = sd.get("byMetro", {}).get(m)
        if st:
            print(f"  {m:22s} median ${st['p50']:>7,}  (n={st['n']:,})")
            samples.append((m, st))
    print("\nSANITY — 5 (role, metro, level) sample medians:")
    shown = 0
    for role in ["Software Developers", "Computer Systems Analysts", "Data Scientists"]:
        bml = roles_out.get(role, {}).get("byMetroLevel", {})
        for k in list(bml)[:2]:
            st = bml[k]
            print(f"  {role} @ {k}: median ${st['p50']:,} (n={st['n']:,})")
            shown += 1
            if shown >= 5:
                break
        if shown >= 5:
            break


if __name__ == "__main__":
    main()
