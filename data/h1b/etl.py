#!/usr/bin/env python3
"""
H-1B Sponsor Finder — ETL  (FEIN-grouped, calamine-accelerated)
===============================================================
Ingests one or more DOL OFLC LCA disclosure XLSX files, keeps only certified
H-1B cases, annualizes wages, and emits a compact (employer, soc_code, state)
rollup that powers the sponsor-finder API.

Source files: https://www.dol.gov/agencies/eta/foreign-labor/performance
              -> "LCA Programs (H-1B, H-1B1, E-3)" disclosure file + record layout

Two things changed from the original streaming ETL:

  1. EMPLOYER GROUPING IS BY FEIN.  The current public LCA files populate
     EMPLOYER_FEIN on ~100% of rows, so we key the rollup on the federal tax id
     instead of a fuzzy name. That collapses "GOOGLE LLC" / "Google Llc" /
     "GOOGLE, LLC." into one employer with no alias map. The display name is the
     most common EMPLOYER_NAME seen for that FEIN.

  2. READING IS VIA python-calamine when available (~10x faster than openpyxl on
     the 100-600 MB files), with the openpyxl read_only streamer kept as a
     dependency-free fallback. Both paths apply the identical
     CASE_STATUS == "Certified" AND VISA_CLASS == "H-1B" filter and yield the
     same cleaned fields.

Usage:
  python etl.py --input LCA_FY2026_Q2.xlsx [LCA_FY2025.xlsx ...] \
                --as-of 2026-03-31 --out sponsors.csv [--limit N] [--sqlite sponsors.db]

  --as-of   anchors the 12-month windows (defaults to today). "last 12 mo" is
            (as_of-365, as_of]; the prior window (as_of-730, as_of-365] powers
            the trend, so feed two fiscal years for a real trend. With a single
            ~6-month file every prev_year_count is 0 (the UI hides the trend
            badge in that case).
  --limit   stop after N kept certified-H-1B rows — fast smoke test on the big
            file without processing all ~half-million rows.

Honest caveats baked into the data:
  * An LCA is an attestation filed BEFORE the USCIS petition. Certification is
    not a petition approval, let alone a hire. Counts are a sponsorship signal.
  * One LCA can cover multiple workers (TOTAL_WORKER_POSITIONS).
"""
import argparse
import csv
import json
import re
import sqlite3
import statistics
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

# --- Column mapping -------------------------------------------------------
# Logical field -> candidate header names (DOL renames columns across years;
# the first header that exists in the file wins).
COLS = {
    "case_status":   ["CASE_STATUS"],
    "visa_class":    ["VISA_CLASS"],
    "decision_date": ["DECISION_DATE"],
    "employer":      ["EMPLOYER_NAME"],
    "fein":          ["EMPLOYER_FEIN", "FEIN"],
    "soc_code":      ["SOC_CODE", "SOC_CD"],
    "soc_title":     ["SOC_TITLE", "SOC_NAME"],
    "state":         ["WORKSITE_STATE", "WORKSITE_STATE_1", "WORKLOC1_STATE"],
    "wage_from":     ["WAGE_RATE_OF_PAY_FROM", "WAGE_RATE_OF_PAY_FROM_1", "WAGE_RATE_FROM"],
    "wage_unit":     ["WAGE_UNIT_OF_PAY", "WAGE_UNIT_OF_PAY_1", "WAGE_RATE_UNIT"],
    "pw_level":      ["PW_WAGE_LEVEL", "PW_WAGE_LEVEL_1", "PW_LEVEL"],
    "total_workers": ["TOTAL_WORKER_POSITIONS", "TOTAL_WORKERS"],
}

REQUIRED = ("case_status", "visa_class", "employer", "fein", "soc_code", "state")

WAGE_MULT = {"YEAR": 1, "HOUR": 2080, "WEEK": 52, "BI-WEEKLY": 26,
             "BIWEEKLY": 26, "MONTH": 12, "HR": 2080, "YR": 1, "MONTHLY": 12,
             "WEEKLY": 52, "HOURLY": 2080}

WS = re.compile(r"\s+")
LEVELS = ("I", "II", "III", "IV")


def clean_name(name):
    """Light display normalization — collapse whitespace, keep original casing."""
    if not name:
        return ""
    return WS.sub(" ", str(name).strip())


def norm_fein(fein):
    """FEIN to a stable key: digits only (drops the NN-NNNNNNN hyphen / spaces)."""
    if fein is None:
        return ""
    return re.sub(r"\D", "", str(fein))


def norm_soc(code):
    """'15-1252.00' -> '15-1252' (strip the O*NET detail suffix)."""
    s = str(code or "").strip()
    return s.split(".")[0]


def annualize(wage, unit):
    try:
        w = float(str(wage).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None
    u = str(unit or "YEAR").upper().strip()
    mult = WAGE_MULT.get(u.replace(" ", ""), WAGE_MULT.get(u, 1))
    val = w * mult
    return int(val) if 10000 <= val <= 5_000_000 else None  # drop obvious junk


def parse_date(dd):
    if dd is None or dd == "":
        return None
    if isinstance(dd, datetime):
        return dd
    if hasattr(dd, "year") and hasattr(dd, "month"):  # date / calamine date
        try:
            return datetime(dd.year, dd.month, dd.day)
        except (ValueError, TypeError):
            return None
    s = str(dd)[:10]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# --- Header resolution ----------------------------------------------------

def resolve_headers(header_row):
    idx = {h: i for i, h in enumerate(header_row) if h}
    resolved = {}
    for field, candidates in COLS.items():
        for c in candidates:
            if c in idx:
                resolved[field] = idx[c]
                break
    missing = [f for f in REQUIRED if f not in resolved]
    if missing:
        raise SystemExit(f"Required columns not found: {missing}\n"
                         f"Headers seen: {list(idx)[:40]}")
    return resolved


def clean_row(r, h):
    """Filter + clean one raw row -> dict, or None if not a certified H-1B."""
    status = str(r[h["case_status"]] or "").strip()
    visa = str(r[h["visa_class"]] or "").strip()
    if visa != "H-1B" or status != "Certified":   # exact: excludes "Certified - Withdrawn"
        return None
    try:
        workers = int(r[h["total_workers"]]) if h.get("total_workers") is not None and r[h["total_workers"]] else 1
    except (ValueError, TypeError):
        workers = 1
    return {
        "fein": norm_fein(r[h["fein"]]),
        "employer": clean_name(r[h["employer"]]),
        "soc_code": norm_soc(r[h["soc_code"]]),
        "soc_title": clean_name(r[h["soc_title"]]) if "soc_title" in h else "",
        "state": str(r[h["state"]] or "").strip().upper(),
        "wage": annualize(r[h["wage_from"]] if "wage_from" in h else None,
                          r[h["wage_unit"]] if "wage_unit" in h else "YEAR"),
        "pw_level": str(r[h["pw_level"]] or "").strip().upper() if "pw_level" in h else "",
        "decision_date": parse_date(r[h["decision_date"]] if "decision_date" in h else None),
        "workers": workers,
    }


# --- Readers (calamine preferred, openpyxl fallback) ----------------------

def _rows_calamine(path):
    from python_calamine import CalamineWorkbook
    wb = CalamineWorkbook.from_path(str(path))
    sheet = wb.get_sheet_by_index(0)
    # iter_rows streams without materializing the whole sheet.
    it = sheet.iter_rows()
    header = next(it)
    yield header
    for r in it:
        yield r


def _rows_openpyxl(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(values_only=True):
        yield r
    wb.close()


def stream_file(path, limit=None, engine="auto"):
    """Yield cleaned certified-H-1B dict rows. Tries calamine, falls back to
    openpyxl. `limit` caps kept rows for fast smoke tests."""
    use = engine
    if use == "auto":
        try:
            import python_calamine  # noqa: F401
            use = "calamine"
        except ImportError:
            use = "openpyxl"
    reader = _rows_calamine if use == "calamine" else _rows_openpyxl
    print(f"  reading {Path(path).name} via {use} ...")

    gen = reader(path)
    header = next(gen)
    h = resolve_headers(header)
    kept = scanned = fein_present = 0
    for r in gen:
        if r is None:
            continue
        scanned += 1
        row = clean_row(r, h)
        if row is None:
            continue
        kept += 1
        if row["fein"]:
            fein_present += 1
        yield row
        if limit and kept >= limit:
            break
    print(f"  {Path(path).name}: scanned {scanned:,} rows, kept {kept:,} certified H-1B"
          f"  (FEIN populated on {fein_present:,} = "
          f"{(100*fein_present/kept if kept else 0):.1f}% of kept)")


# --- Rollup ---------------------------------------------------------------

class Agg:
    __slots__ = ("count", "workers", "wages", "levels", "names", "titles",
                 "last_filed", "prev")

    def __init__(self):
        self.count = 0
        self.workers = 0
        self.wages = []
        self.levels = Counter()
        self.names = Counter()
        self.titles = Counter()
        self.last_filed = None
        self.prev = 0


def build_rollup(rows, as_of):
    as_of = parse_date(as_of) or datetime.today()
    w1_start = as_of - timedelta(days=365)
    w2_start = as_of - timedelta(days=730)

    groups = defaultdict(Agg)   # (fein, soc_code, state) -> Agg

    for row in rows:
        fein, soc, st = row["fein"], row["soc_code"], row["state"]
        if not fein or not soc or not st:
            continue
        key = (fein, soc, st)
        a = groups[key]
        dd = row["decision_date"]
        in_last12 = dd is not None and w1_start < dd <= as_of
        in_prev12 = dd is not None and w2_start < dd <= w1_start

        if in_prev12:
            a.prev += 1
        # last-12-mo aggregates (the numbers shown on a card). When dates are
        # missing entirely we still count the row so the file isn't empty.
        if in_last12 or dd is None:
            a.count += 1
            a.workers += row["workers"]
            if row["wage"] is not None:
                a.wages.append(row["wage"])
            if row["pw_level"] in LEVELS:
                a.levels[row["pw_level"]] += 1
        if row["employer"]:
            a.names[row["employer"]] += 1
        if row["soc_title"]:
            a.titles[row["soc_title"]] += 1
        if dd is not None and (a.last_filed is None or dd > a.last_filed):
            a.last_filed = dd

    out = []
    for (fein, soc, st), a in groups.items():
        if a.count == 0 and a.prev == 0:
            continue
        employer = a.names.most_common(1)[0][0] if a.names else fein
        soc_title = a.titles.most_common(1)[0][0] if a.titles else ""
        out.append({
            "employer": employer,
            "soc_code": soc,
            "soc_title": soc_title,
            "state": st,
            "lca_count": a.count,
            "worker_positions": a.workers,
            "median_wage": int(statistics.median(a.wages)) if a.wages else "",
            "wage_levels": json.dumps({k: a.levels[k] for k in LEVELS if a.levels[k]}),
            "prev_year_count": a.prev,
            "trend": ("up" if a.count > a.prev else "down" if a.count < a.prev else "flat"),
            "last_filed": a.last_filed.date().isoformat() if a.last_filed else "",
        })

    out.sort(key=lambda d: (d["soc_code"], d["state"], -d["lca_count"]))
    return out


FIELDNAMES = ["employer", "soc_code", "soc_title", "state", "lca_count",
              "worker_positions", "median_wage", "wage_levels",
              "prev_year_count", "trend", "last_filed"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", nargs="+", required=True, help="OFLC LCA xlsx file(s)")
    ap.add_argument("--as-of", default=datetime.today().strftime("%Y-%m-%d"))
    ap.add_argument("--out", default="sponsors.csv")
    ap.add_argument("--limit", type=int, default=None,
                    help="stop after N kept rows (smoke test)")
    ap.add_argument("--engine", choices=["auto", "calamine", "openpyxl"], default="auto")
    ap.add_argument("--sqlite", default=None, help="optional SQLite file for quick local testing")
    ap.add_argument("--aliases", default=None, help="(legacy, unused) FEIN grouping needs no alias map")
    args = ap.parse_args()

    print(f"Streaming {len(args.input)} file(s); grouping by FEIN ...")

    def all_rows():
        for f in args.input:
            yield from stream_file(f, limit=args.limit, engine=args.engine)

    rollup = build_rollup(all_rows(), args.as_of)
    if not rollup:
        raise SystemExit("No certified H-1B rows found — check the input file/columns.")

    unique_employers = len({(r["employer"]) for r in rollup})
    total_lca = sum(r["lca_count"] for r in rollup)
    print(f"Rollup: {len(rollup):,} rows  |  {unique_employers:,} unique employers"
          f"  |  {total_lca:,} certified LCAs counted")

    with open(args.out, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(rollup)
    print(f"Wrote {len(rollup):,} rollup rows -> {args.out}")

    if args.sqlite:
        con = sqlite3.connect(args.sqlite)
        con.execute("DROP TABLE IF EXISTS sponsors")
        con.execute("""CREATE TABLE sponsors (employer TEXT, soc_code TEXT, soc_title TEXT,
                       state TEXT, lca_count INT, worker_positions INT, median_wage INT,
                       wage_levels TEXT, prev_year_count INT, trend TEXT, last_filed TEXT)""")
        con.executemany(
            "INSERT INTO sponsors VALUES (:employer,:soc_code,:soc_title,:state,:lca_count,"
            ":worker_positions,:median_wage,:wage_levels,:prev_year_count,:trend,:last_filed)",
            [{**r, "median_wage": r["median_wage"] or None} for r in rollup])
        con.execute("CREATE INDEX idx_soc_state ON sponsors(soc_code, state)")
        con.commit(); con.close()
        print(f"Wrote SQLite -> {args.sqlite}")

    return rollup


if __name__ == "__main__":
    main()
